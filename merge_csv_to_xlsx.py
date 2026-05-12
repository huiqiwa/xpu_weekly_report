import ast
import argparse
import glob
import itertools
import json
import os
import re

import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

_table_counter = 0


def add_excel_table(worksheet, start_row, num_rows, num_cols, table_name_prefix="Tbl"):
    """Apply Excel Table formatting (like Ctrl+T) to a data range."""
    global _table_counter
    if num_rows == 0 or num_cols == 0:
        return
    _table_counter += 1
    end_col_letter = get_column_letter(num_cols)
    ref = f"A{start_row + 1}:{end_col_letter}{start_row + 1 + num_rows}"
    table_name = re.sub(r"[^A-Za-z0-9_.]", "_", f"{table_name_prefix}_{_table_counter}")
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    table.tableStyleInfo = style
    worksheet.add_table(table)


def auto_fit_columns(worksheet, max_width=80):
    """Auto-adjust column widths based on cell content."""
    for col_cells in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        worksheet.column_dimensions[col_letter].width = min(max_len + 4, max_width)


_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PARENT_DIR = os.path.dirname(_SCRIPT_DIR)
_REPORTS_ROOT = os.path.join(_SCRIPT_DIR, "reports")
_report_candidates = sorted(
    [d for d in glob.glob(os.path.join(glob.escape(_REPORTS_ROOT), "reports_*")) if os.path.isdir(d)]
)
if not _report_candidates:
    raise FileNotFoundError(f"No report directories found under: {_REPORTS_ROOT}")
BASE_DIR = _report_candidates[-1]
WORKLOADS_DIR = os.path.join(_PARENT_DIR, "xpu-perf", "projects", "micro_perf", "workloads")

OP_GROUPS = [
    (
        "Norm & Quant",
        [
            "scale_dynamic_quant",
            "head_rms_norm",
            "head_rms_norm_dynamic_quant",
            "add_rms_norm_dynamic_quant",
        ],
    ),
    (
        "Attention & rope & kvcache",
        [
            "rotary_embedding",
            "store_kv_cache",
            "dequant_kv_cache",
            "flash_attention",
        ],
    ),
    (
        "gemm & group_gemm & moe_ops",
        [
            "moe_gating_gemm",
            "quant_matmul",
            "moe_quant_group_gemm",
            "moe_softmax_topk",
            "moe_scatter_dynamic_quant",
            "moe_swiglu_dynamic_quant",
            "swiglu_dynamic_quant",
            "moe_gather",
            "moe_quant_group_gemm_combine",
            "quant_group_gemm_reduce_sum",
        ],
    ),
    ("tensor_gemm_ops", ["gemm"]),
    ("vector_activation_ops", ["gelu", "silu"]),
    (
        "vector_index_ops",
        ["embedding", "gather", "index_add", "index_select", "scatter"],
    ),
    ("vector_linear_ops", ["add", "cast", "mul", "sub"]),
    ("vector_norm_ops", ["layer_norm", "rms_norm", "softmax"]),
    ("vector_reduction_ops", ["reduce_max", "reduce_min", "reduce_sum", "topk"]),
    ("vector_sfu_ops", ["cos", "div", "exp", "log", "sin", "sqrt"]),
    (
        "xccl ops",
        [
            "all_reduce",
            "all_gather",
            "reduce_scatter",
            "all_to_all",
            "device2device",
            "device2host",
            "host2device",
        ],
    ),
    (
        "sage_attention",
        [
            "sage_attention_page",
            "sage_attention_decode_page",
            "sage_attention_v1",
        ],
    ),
]

ORDERED_OPS = [op_name for _, op_names in OP_GROUPS for op_name in op_names]
OP_CATEGORY_MAP = {
    op_name: category_name for category_name, op_names in OP_GROUPS for op_name in op_names
}
OP_ORDER_MAP = {op_name: index for index, op_name in enumerate(ORDERED_OPS)}

PEAK_BW_GBS = 456.0
PEAK_TFLOPS_FP32 = 12.28
PEAK_TFLOPS_LOW = 98.5  # bf16 / fp16
PEAK_TFLOPS_INT8 = 197.0
METRIC_COLUMNS = {
    "sku_name",
    "op_name",
    "provider",
    "latency(us)",
    "read_bytes(B)",
    "write_bytes(B)",
    "io_bytes(B)",
    "mem_bw(GB/s)",
    "calc_flops",
    "calc_flops_power(tflops)",
    "calc_mem_ratio",
    "kernels",
    "MFU(%)",
    "MBU(%)",
}
CASE_META_COLUMNS = {"sku_name", "op_name", "provider", "_source_json"}


def find_csv_files(base_dir):
    pattern = os.path.join(glob.escape(base_dir), "*", "*", "*.csv")
    return sorted(glob.glob(pattern))


def resolve_report_csv_base(base_dir):
    intel_root = os.path.join(base_dir, "INTEL")
    candidates = [path for path in sorted(glob.glob(os.path.join(glob.escape(intel_root), "*"))) if os.path.isdir(path)]
    if not candidates:
        raise FileNotFoundError(f"No device report directory found under: {intel_root}")
    if len(candidates) > 1:
        raise ValueError(
            "Multiple device report directories found under INTEL; please keep only one target device directory."
        )
    return candidates[0]


def normalize_scalar(value):
    if pd.isna(value):
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return ("{:.12g}".format(value)).lower()
    text = str(value).strip()
    lowered = text.lower()
    if lowered in {"true", "false"}:
        return lowered
    try:
        number = float(text)
    except ValueError:
        return lowered
    if number.is_integer():
        return str(int(number))
    return ("{:.12g}".format(number)).lower()


SAGE_ATTENTION_OPS = {"sage_attention_page", "sage_attention_decode_page", "sage_attention_v1"}


def compute_mfu(row):
    op_name = str(row.get("op_name", ""))
    if op_name in SAGE_ATTENTION_OPS:
        peak = PEAK_TFLOPS_INT8
    else:
        dtype = str(row.get("dtype", "")).lower()
        if "float32" in dtype or dtype in ("fp32", "tf32", "tfloat32"):
            peak = PEAK_TFLOPS_FP32
        elif dtype == "int8":
            peak = PEAK_TFLOPS_INT8
        else:
            peak = PEAK_TFLOPS_LOW
    value = row.get("calc_flops_power(tflops)")
    if pd.isna(value) or peak == 0:
        return None
    return round(value / peak * 100, 4)


def compute_mbu(row):
    value = row.get("mem_bw(GB/s)")
    if pd.isna(value):
        return None
    return round(value / PEAK_BW_GBS * 100, 4)


def build_case_signature(case_dict):
    comparable = []
    for key in sorted(case_dict):
        if key.startswith("_") or key in METRIC_COLUMNS or key in CASE_META_COLUMNS:
            continue
        comparable.append((key, normalize_scalar(case_dict[key])))
    return tuple(comparable)


def value_options(key, value):
    parts = key.split(".")
    if len(parts) > 1:
        if isinstance(value, list):
            if value and isinstance(value[0], list):
                return [dict(zip(parts, item)) for item in value]
            if len(value) == len(parts):
                return [dict(zip(parts, value))]
        return [dict(zip(parts, [value]))]

    if isinstance(value, list):
        return [{key: item} for item in value]
    return [{key: value}]


def expand_case(case_def):
    option_groups = []
    for key, value in case_def.items():
        option_groups.append(value_options(key, value))

    expanded = []
    for combination in itertools.product(*option_groups):
        merged = {}
        for item in combination:
            merged.update(item)
        expanded.append(merged)
    return expanded


def load_workload_cases(workloads_dir):
    workload_cases = {}
    json_files = sorted(
        glob.glob(os.path.join(glob.escape(workloads_dir), "**", "*.json"), recursive=True)
    )

    for json_path in json_files:
        with open(json_path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)

        if isinstance(payload, dict) and "cases" in payload:
            op_name = os.path.splitext(os.path.basename(json_path))[0]
            case_defs = payload["cases"]
            source_name = os.path.relpath(json_path, workloads_dir)
            workload_cases.setdefault(op_name, [])
            for case_def in case_defs:
                for expanded_case in expand_case(case_def):
                    expanded_case["_source_json"] = source_name
                    workload_cases[op_name].append(expanded_case)
            continue

        if not isinstance(payload, dict):
            continue

        source_name = os.path.relpath(json_path, workloads_dir)
        for op_name, case_defs in payload.items():
            if not isinstance(case_defs, list):
                continue
            workload_cases.setdefault(op_name, [])
            for case_def in case_defs:
                for expanded_case in expand_case(case_def):
                    expanded_case["_source_json"] = source_name
                    workload_cases[op_name].append(expanded_case)

    return workload_cases


def filter_workload_cases(workload_cases):
    return {op_name: workload_cases[op_name] for op_name in ORDERED_OPS if op_name in workload_cases}


def load_report_frames(base_dir):
    grouped_frames = {}  # {op_name: {provider: [df, ...]}}
    for csv_path in find_csv_files(base_dir):
        rel_path = os.path.relpath(csv_path, base_dir)
        parts = rel_path.split(os.sep)
        op_name = parts[0]
        provider = parts[1] if len(parts) > 1 else "unknown"
        grouped_frames.setdefault(op_name, {}).setdefault(provider, []).append(pd.read_csv(csv_path))

    merged_frames = {}  # {op_name: {provider: df}}
    for op_name, provider_dict in grouped_frames.items():
        merged_frames[op_name] = {}
        for provider, frames in provider_dict.items():
            merged = pd.concat(frames, ignore_index=True, sort=False)
            merged["MFU(%)"] = merged.apply(compute_mfu, axis=1)
            merged["MBU(%)"] = merged.apply(compute_mbu, axis=1)
            merged_frames[op_name][provider] = merged
    return merged_frames


def filter_report_frames(report_frames):
    return {op_name: report_frames[op_name] for op_name in ORDERED_OPS if op_name in report_frames}


def case_matches_row(expected_case, row, report_columns):
    comparable_columns = [
        column
        for column in expected_case
        if not column.startswith("_") and column in report_columns and column not in METRIC_COLUMNS
    ]
    if not comparable_columns:
        return False

    for column in comparable_columns:
        if normalize_scalar(expected_case[column]) != normalize_scalar(row[column]):
            return False
    return True


def append_missing_cases_to_reports(workload_cases, report_frames, registered_df, sku_name=""):
    enriched_frames = {}  # {op_name: {provider: df}}

    # Build a mapping of op_name -> list of registered providers
    op_providers_map = {}
    if not registered_df.empty:
        for op_name, provider_group in registered_df.groupby("op_name"):
            op_providers_map[op_name] = sorted(provider_group["provider"].dropna().astype(str).unique())

    all_ops = sorted(set(report_frames) | set(workload_cases))

    for op_name in all_ops:
        provider_dict = report_frames.get(op_name, {})
        registered_providers = op_providers_map.get(op_name, [])
        all_providers = sorted(set(provider_dict.keys()) | set(registered_providers))
        expected_cases = workload_cases.get(op_name, [])

        # Fallback: if no providers known but workload cases exist, use "torch" as default
        if not all_providers and expected_cases:
            all_providers = ["torch"]

        enriched_frames[op_name] = {}

        for provider in all_providers:
            report_df = provider_dict.get(provider)
            if report_df is not None:
                working_df = report_df.copy()
            else:
                working_df = pd.DataFrame()

            if not working_df.empty and "case_status" not in working_df.columns:
                working_df["case_status"] = "SUCCESS"

            report_columns = set(working_df.columns)
            missing_rows = []

            # Pre-normalize report rows for O(1) lookup instead of O(n) per case
            _norm_rows = []
            if not working_df.empty:
                _comparable_cols_all = [c for c in working_df.columns
                                        if c not in METRIC_COLUMNS and not c.startswith("_")]
                for _, row in working_df.iterrows():
                    _norm_rows.append({c: normalize_scalar(row[c]) for c in _comparable_cols_all})
            _sig_cache = {}  # tuple of comparable columns -> set of value tuples

            for expected_case in expected_cases:
                comparable_columns = tuple(sorted(
                    c for c in expected_case
                    if not c.startswith("_") and c in report_columns and c not in METRIC_COLUMNS
                ))
                if not comparable_columns:
                    continue

                if comparable_columns not in _sig_cache:
                    sigs = set()
                    for nr in _norm_rows:
                        sigs.add(tuple(nr.get(c, "") for c in comparable_columns))
                    _sig_cache[comparable_columns] = sigs

                expected_sig = tuple(normalize_scalar(expected_case[c]) for c in comparable_columns)
                if expected_sig in _sig_cache[comparable_columns]:
                    continue

                missing_row = {
                    "sku_name": sku_name,
                    "op_name": op_name,
                    "provider": provider,
                    "case_status": "FAILED",
                }
                for key, value in expected_case.items():
                    if key.startswith("_"):
                        continue
                    if not report_columns or key in report_columns:
                        missing_row[key] = value
                for metric_column in METRIC_COLUMNS:
                    if not report_columns or metric_column in report_columns:
                        missing_row.setdefault(metric_column, None)
                missing_rows.append(missing_row)

            if missing_rows:
                missing_df = pd.DataFrame(missing_rows)
                combined_df = pd.concat([working_df, missing_df], ignore_index=True, sort=False)
            else:
                combined_df = working_df

            if combined_df.empty:
                continue

            if "source_json" in combined_df.columns:
                combined_df = combined_df.drop(columns=["source_json"])

            combined_df = combined_df.dropna(axis=1, how="all")

            ordered_columns = list(combined_df.columns)
            if "case_status" in ordered_columns:
                ordered_columns.remove("case_status")
                if "MFU(%)" in ordered_columns:
                    insert_index = ordered_columns.index("MFU(%)")
                    ordered_columns.insert(insert_index, "case_status")
                else:
                    ordered_columns.append("case_status")
            enriched_frames[op_name][provider] = combined_df[ordered_columns]

    return enriched_frames


def parse_registered_providers(log_text, log_name):
    rows = []
    current_provider = None
    in_op_table = False

    for raw_line in log_text.splitlines():
        line = raw_line.strip()
        if line.startswith("Provider:"):
            current_provider = line.split(":", 1)[1].strip()
            in_op_table = False
            continue
        if not current_provider or not line.startswith("|"):
            continue

        parts = [part.strip() for part in line.split("|") if part.strip()]
        if len(parts) < 2:
            continue
        if parts[0] == "op_name" and parts[1] == "op_cls":
            in_op_table = True
            continue
        if parts[0] in {"op_name", "attr", "provider", "key", "value"}:
            continue
        if parts[0].startswith("+"):
            continue
        if in_op_table and len(parts) >= 2 and parts[1].startswith("<class ") and re.match(r"^[A-Za-z0-9_]+$", parts[0]):
            rows.append(
                {
                    "op_name": parts[0],
                    "provider": current_provider,
                    "source_log": log_name,
                }
            )

    return rows


def parse_failed_cases(log_text, log_name):
    rows = []
    pattern = re.compile(
        r"Failed to create op (?P<op_name>\S+) with provider (?P<provider>\S+) with args (?P<args>\{.*?\}) with error (?P<error>.*)"
    )

    for match in pattern.finditer(log_text):
        args_text = match.group("args")
        try:
            case_args = ast.literal_eval(args_text)
        except (ValueError, SyntaxError):
            case_args = {"raw_args": args_text}

        row = {
            "op_name": match.group("op_name"),
            "provider": match.group("provider"),
            "error": match.group("error").strip(),
            "source_log": log_name,
        }
        row.update(case_args)
        rows.append(row)

    return rows


def load_launch_log_data(logs_dir):
    log_paths = sorted(glob.glob(os.path.join(glob.escape(logs_dir), "*.txt")))
    registered_rows = []
    failed_rows = []

    for log_path in log_paths:
        with open(log_path, "r", encoding="utf-8", errors="ignore") as handle:
            log_text = handle.read()

        log_name = os.path.basename(log_path)
        registered_rows.extend(parse_registered_providers(log_text, log_name))
        failed_rows.extend(parse_failed_cases(log_text, log_name))

    registered_df = pd.DataFrame(registered_rows).drop_duplicates() if registered_rows else pd.DataFrame()
    failed_df = pd.DataFrame(failed_rows) if failed_rows else pd.DataFrame()
    if not failed_df.empty:
        failed_df = failed_df.drop_duplicates()
    return registered_df, failed_df


def build_report_provider_stats(report_frames):
    stats = []
    for op_name, provider_dict in report_frames.items():
        for provider, report_df in provider_dict.items():
            unique_cases = {
                build_case_signature(record)
                for record in report_df.to_dict("records")
            }
            stats.append(
                {
                    "op_name": op_name,
                    "provider": str(provider),
                    "report_rows": len(report_df),
                    "report_case_count": len(unique_cases),
                }
            )
    return pd.DataFrame(stats)


def build_provider_status(report_frames, registered_df, failed_df, enriched_report_frames=None):
    report_stats_df = build_report_provider_stats(report_frames)
    frames = [frame for frame in [report_stats_df, registered_df, failed_df] if not frame.empty]
    if frames:
        universe_df = pd.concat(
            [frame[["op_name", "provider"]] for frame in frames],
            ignore_index=True,
        ).drop_duplicates()
    else:
        universe_df = pd.DataFrame(columns=["op_name", "provider"])

    missing_ops = [op_name for op_name in ORDERED_OPS if op_name not in set(universe_df.get("op_name", []))]
    if missing_ops:
        universe_df = pd.concat(
            [
                universe_df,
                pd.DataFrame({"op_name": missing_ops, "provider": [""] * len(missing_ops)}),
            ],
            ignore_index=True,
        ).drop_duplicates()

    # Include providers from enriched report frames
    if enriched_report_frames:
        enriched_pairs = [
            {"op_name": op_name, "provider": provider}
            for op_name, provider_dict in enriched_report_frames.items()
            for provider in provider_dict
        ]
        if enriched_pairs:
            universe_df = pd.concat(
                [universe_df, pd.DataFrame(enriched_pairs)],
                ignore_index=True,
            ).drop_duplicates()

    # Remove empty-provider placeholder rows if the op already has real provider rows
    ops_with_providers = set(universe_df.loc[universe_df["provider"].fillna("").astype(str).str.strip() != "", "op_name"])
    universe_df = universe_df[
        (universe_df["provider"].fillna("").astype(str).str.strip() != "")
        | (~universe_df["op_name"].isin(ops_with_providers))
    ].reset_index(drop=True)

    registered_pairs = set()
    if not registered_df.empty:
        registered_pairs = set(map(tuple, registered_df[["op_name", "provider"]].itertuples(index=False, name=None)))

    report_stats = {}
    if not report_stats_df.empty:
        report_stats = {
            (row.op_name, row.provider): {
                "report_rows": int(row.report_rows),
                "report_case_count": int(row.report_case_count),
            }
            for row in report_stats_df.itertuples(index=False)
        }

    failure_groups = {}
    if not failed_df.empty:
        for (op_name, provider), group in failed_df.groupby(["op_name", "provider"]):
            unique_cases = {
                build_case_signature(record)
                for record in group.to_dict("records")
            }
            errors = [error for error in group["error"].fillna("").astype(str).unique() if error]
            failure_groups[(op_name, provider)] = {
                "failed_case_count": len(unique_cases),
                "error_msg": " | ".join(errors[:3]),
            }

    # Supplement failure_groups with FAILED case counts from enriched report frames
    if enriched_report_frames:
        for op_name, provider_dict in enriched_report_frames.items():
            for provider, edf in provider_dict.items():
                if "case_status" not in edf.columns:
                    continue
                enriched_failed = int((edf["case_status"] == "FAILED").sum())
                if enriched_failed == 0:
                    continue
                key = (op_name, provider)
                existing = failure_groups.get(key, {"failed_case_count": 0, "error_msg": ""})
                if enriched_failed > existing["failed_case_count"]:
                    failure_groups[key] = {
                        "failed_case_count": enriched_failed,
                        "error_msg": existing["error_msg"],
                    }

    # Build total case counts from enriched report frames
    total_cases = {}
    if enriched_report_frames:
        for op_name, provider_dict in enriched_report_frames.items():
            for provider, edf in provider_dict.items():
                total_cases[(op_name, provider)] = len(edf)

    status_rows = []
    for row in universe_df.itertuples(index=False):
        key = (row.op_name, row.provider)
        failure_info = failure_groups.get(key, {"failed_case_count": 0, "error_msg": ""})
        total = total_cases.get(key, 0)
        failed = failure_info["failed_case_count"]

        status_rows.append(
            {
                "op_name": row.op_name,
                "provider": row.provider,
                "failed_case_count": f"{failed}/{total}",
                "error_msg": failure_info["error_msg"],
            }
        )

    status_df = pd.DataFrame(status_rows)
    if not status_df.empty:
        status_df["category"] = status_df["op_name"].map(OP_CATEGORY_MAP)
        status_df["_op_order"] = status_df["op_name"].map(OP_ORDER_MAP)
        status_df = status_df[status_df["op_name"].isin(ORDERED_OPS)]
        status_df = status_df.sort_values(["_op_order", "provider"]).reset_index(drop=True)
        status_df = status_df[["category", "op_name", "provider", "failed_case_count"]]
    return status_df


def parse_args():
    parser = argparse.ArgumentParser(description="Merge xpu-perf CSV reports into one xlsx workbook.")
    parser.add_argument("--workloads-dir", default=WORKLOADS_DIR, help="Workloads JSON directory.")
    return parser.parse_args()


def main():
    args = parse_args()
    base_dir = BASE_DIR
    workloads_dir = args.workloads_dir
    logs_dir = base_dir
    script_dir = os.path.dirname(os.path.abspath(__file__))
    reports_dir = os.path.join(script_dir, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    output_file = os.path.join(reports_dir, os.path.basename(base_dir) + ".xlsx")
    report_csv_base = resolve_report_csv_base(base_dir)
    sku_name = os.path.basename(report_csv_base)

    report_frames = load_report_frames(report_csv_base)
    report_frames = filter_report_frames(report_frames)
    workload_cases = filter_workload_cases(load_workload_cases(workloads_dir))
    registered_df, failed_df = load_launch_log_data(logs_dir)
    enriched_report_frames = append_missing_cases_to_reports(workload_cases, report_frames, registered_df, sku_name)
    provider_status_df = build_provider_status(report_frames, registered_df, failed_df, enriched_report_frames)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        if provider_status_df.empty:
            provider_status_df = pd.DataFrame([{"category": "", "op_name": "", "provider": "", "failed_case_count": 0}])
        provider_status_df["sheet_link"] = ""
        provider_status_df["AICE Owner"] = ""
        provider_status_df["Next step Plan"] = ""
        provider_status_df.to_excel(writer, sheet_name="Summary", index=False)
        summary_ws = writer.sheets["Summary"]
        # Build sheet_link column hyperlinks (column index for sheet_link)
        link_col_idx = list(provider_status_df.columns).index("sheet_link") + 1  # 1-based
        link_col_letter = get_column_letter(link_col_idx)
        op_sheet_names = {}  # op_name -> sheet_name mapping, filled later

        for op_name in ORDERED_OPS:
            if op_name not in enriched_report_frames:
                continue
            provider_dict = enriched_report_frames[op_name]
            if not provider_dict:
                pd.DataFrame().to_excel(writer, sheet_name=op_name[:31], index=False)
                print(f"  Added sheet: {op_name[:31]} (0 rows)")
                op_sheet_names[op_name] = op_name[:31]
                continue

            sheet_name = op_name[:31]
            op_sheet_names[op_name] = sheet_name
            providers = sorted(provider_dict.keys())
            current_row = 1  # row 0 reserved for back link
            total_rows = 0
            missing_count = 0

            for provider in providers:
                provider_df = provider_dict[provider]
                if "case_status" in provider_df.columns:
                    missing_count += int((provider_df["case_status"] == "FAILED").sum())
                provider_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=current_row)
                ws = writer.sheets[sheet_name]
                add_excel_table(ws, current_row, len(provider_df), len(provider_df.columns), f"{sheet_name}_{provider}")
                total_rows += len(provider_df)
                current_row += len(provider_df) + 2  # header + data + 1 blank row

            # Add "Back to Summary" hyperlink in cell A1 of each op sheet
            ws = writer.sheets[sheet_name]
            ws["A1"] = "← Back to Summary"
            ws["A1"].hyperlink = "#Summary!A1"
            ws["A1"].style = "Hyperlink"

            print(
                f"  Added sheet: {sheet_name} ({total_rows} rows, providers: {', '.join(providers)}, missing_cases: {missing_count})"
            )
            auto_fit_columns(writer.sheets[sheet_name])

        # Now add hyperlinks in Summary sheet for each op row
        for row_idx, row_data in provider_status_df.iterrows():
            excel_row = row_idx + 2  # 1 for header, 1 for 0-based index
            target_op = row_data["op_name"]
            if target_op in op_sheet_names:
                target_sheet = op_sheet_names[target_op]
                cell = summary_ws[f"{link_col_letter}{excel_row}"]
                cell.value = target_sheet
                cell.hyperlink = f"#{target_sheet}!A1"
                cell.style = "Hyperlink"
        add_excel_table(summary_ws, 0, len(provider_status_df), len(provider_status_df.columns), "Summary")
        auto_fit_columns(summary_ws)
        # Fix "Next step Plan" column width
        plan_col_idx = list(provider_status_df.columns).index("Next step Plan") + 1
        summary_ws.column_dimensions[get_column_letter(plan_col_idx)].width = 47
        print(f"  Added sheet: Summary ({len(provider_status_df)} rows)")

    print(f"\nDone! Output: {output_file}")


if __name__ == "__main__":
    main()
