#!/usr/bin/env python3
"""Compare per-case latency between two report directories, output one XLSX.

Usage:
    python compare_latency_xlsx.py <report_dir_old> <report_dir_new> [-o output.xlsx]

Each sheet corresponds to one (op, provider) combination.
Columns: config columns | old_latency(us) | new_latency(us) | ratio(old/new) | change
  ratio > 1 means new is slower, ratio < 1 means new is faster.
"""

import argparse
import csv
import glob
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Result / metric columns that should NOT be used as case keys.
_METRIC_COLUMNS = {
    "sku_name", "op_name", "provider",
    "latency(us)", "mem_bw(GB/s)", "calc_flops_power(tflops)",
    "read_bytes(B)", "write_bytes(B)", "io_bytes(B)",
    "calc_flops", "calc_mem_ratio", "kernels",
    "MFU(%)", "MBU(%)", "case_status",
    # CCL-specific
    "algo_size(B)", "bus_size(B)",
    "algo_bw(GB/s)", "bus_bw(GB/s)",
    "algo_bw_sum(GB/s)", "bus_bw_sum(GB/s)",
    "latency_list(us)", "algo_bw_list(GB/s)", "bus_bw_list(GB/s)",
}

# Colours
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_OLD_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_NEW_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def find_intel_base(report_dir):
    """Find the INTEL/<gpu>/ or GPU/<gpu>/ subdirectory."""
    for vendor in ("INTEL", "GPU"):
        vendor_path = os.path.join(report_dir, vendor)
        if not os.path.isdir(vendor_path):
            continue
        for name in sorted(os.listdir(vendor_path)):
            candidate = os.path.join(vendor_path, name)
            if os.path.isdir(candidate):
                return candidate
    return None


def load_csvs(base_dir):
    """Return {(op, provider): [rows]} where rows are dicts from csv.DictReader."""
    result = {}
    for root, _dirs, files in os.walk(base_dir):
        for f in files:
            if not f.endswith(".csv"):
                continue
            full = os.path.join(root, f)
            rel = os.path.relpath(full, base_dir)
            parts = rel.split(os.sep)
            if len(parts) < 3:
                continue
            op, provider = parts[0], parts[1]
            if provider == "ipex":
                continue
            with open(full) as fh:
                rows = list(csv.DictReader(fh))
            result.setdefault((op, provider), []).extend(rows)
    return result


def config_keys(row):
    """Return ordered list of config column names (non-metric)."""
    return [k for k in row.keys() if k not in _METRIC_COLUMNS]


def make_case_key(row, key_cols):
    """Build a hashable tuple from the config columns of a row."""
    return tuple(row.get(k, "") for k in key_cols)


# ---------------------------------------------------------------------------
# Timestamp / label helpers
# ---------------------------------------------------------------------------

def extract_timestamp(dirname):
    m = re.search(r"(\d{4}(-\d{2}){5})", os.path.basename(dirname))
    return m.group(1).replace("-", "") if m else None


def make_label(dirname):
    m = re.search(r"\d{4}-(\d{2}-\d{2})-\d{2}-\d{2}-\d{2}", os.path.basename(dirname))
    return m.group(1) if m else os.path.basename(dirname)[:10]


# ---------------------------------------------------------------------------
# Sheet name helper — Excel limits to 31 chars, no special chars
# ---------------------------------------------------------------------------

def safe_sheet_name(name, max_len=31):
    # Replace characters invalid in Excel sheet names
    for ch in "\\/*?:[]":
        name = name.replace(ch, "_")
    if len(name) > max_len:
        name = name[:max_len]
    return name


# ---------------------------------------------------------------------------
# Op ordering (matches run_test.sh execution order)
# ---------------------------------------------------------------------------

ORDERED_OPS = [
    # Norm & Quant
    "scale_dynamic_quant",
    "head_rms_norm",
    "head_rms_norm_dynamic_quant",
    "add_rms_norm_dynamic_quant",
    # Attention & rope & kvcache
    "rotary_embedding",
    "store_kv_cache",
    "dequant_kv_cache",
    "flash_attention",
    # gemm & group_gemm & moe_ops
    "moe_gating_gemm",
    "quant_matmul",
    "moe_quant_group_gemm",
    "moe_softmax_topk",
    "moe_scatter_dynamic_quant",
    "moe_swiglu_dynamic_quant",
    "swiglu_dynamic_quant",
    "moe_gather",
    "moe_quant_group_gemm_combine",
    "quant_group_gemm_reduce_sum",
    # tensor_gemm_ops
    "gemm",
    # vector_activation_ops
    "gelu",
    "silu",
    # vector_index_ops
    "embedding",
    "gather",
    "index_add",
    "index_select",
    "scatter",
    # vector_linear_ops
    "add",
    "cast",
    "mul",
    "sub",
    # vector_norm_ops
    "layer_norm",
    "rms_norm",
    "softmax",
    # vector_reduction_ops
    "reduce_max",
    "reduce_min",
    "reduce_sum",
    "topk",
    # vector_sfu_ops
    "cos",
    "div",
    "exp",
    "log",
    "sin",
    "sqrt",
    # sage_attention
    "sage_attention_page",
    "sage_attention_decode_page",
    "sage_attention_v1",
    # xccl ops
    "device2device",
    "device2host",
    "host2device",
    "all_reduce",
    "all_gather",
    "reduce_scatter",
    "all_to_all",
]

_OP_ORDER = {op: i for i, op in enumerate(ORDERED_OPS)}


def _op_sort_key(key):
    """Sort (op, provider) by ORDERED_OPS position, then provider name."""
    op, provider = key
    return (_OP_ORDER.get(op, len(ORDERED_OPS)), op, provider)


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

def write_xlsx(old_data, new_data, old_label, new_label, output_path):
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Gather all (op, provider) keys and sort by run_test.sh order
    all_keys = sorted(set(old_data.keys()) | set(new_data.keys()), key=_op_sort_key)

    sheet_name_counts = {}
    op_sheet_map = {}  # (op, provider) -> sheet_name

    for op, provider in all_keys:
        old_rows = old_data.get((op, provider), [])
        new_rows = new_data.get((op, provider), [])

        if not old_rows and not new_rows:
            continue

        # Determine config columns from whichever has data
        sample = old_rows[0] if old_rows else new_rows[0]
        key_cols = config_keys(sample)

        # Also check new side for any additional config columns
        if new_rows:
            new_key_cols = config_keys(new_rows[0])
            # Use union preserving order
            for k in new_key_cols:
                if k not in key_cols:
                    key_cols.append(k)

        # Build maps: case_key -> row
        old_map = {}
        for r in old_rows:
            k = make_case_key(r, key_cols)
            old_map[k] = r
        new_map = {}
        for r in new_rows:
            k = make_case_key(r, key_cols)
            new_map[k] = r

        all_cases = sorted(set(old_map.keys()) | set(new_map.keys()))
        if not all_cases:
            continue

        # Create sheet
        raw_name = f"{op}-{provider}"
        sname = safe_sheet_name(raw_name)
        # Handle duplicates
        if sname in sheet_name_counts:
            sheet_name_counts[sname] += 1
            sname = safe_sheet_name(f"{raw_name}_{sheet_name_counts[sname]}")
        else:
            sheet_name_counts[sname] = 0

        ws = wb.create_sheet(title=sname)
        op_sheet_map[(op, provider)] = sname

        # "← Back to Summary" hyperlink in row 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        back_cell = ws.cell(row=1, column=1, value="← Back to Summary")
        back_cell.hyperlink = "#'Summary'!A1"
        back_cell.font = Font(color="0563C1", underline="single")

        # --- Header row (row 2) ---
        headers = (
            key_cols
            + [f"latency_old({old_label})(us)", f"latency_new({new_label})(us)",
               "ratio(new/old)", "change"]
        )
        header_row = 2
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = _THIN_BORDER

        # Column indices for latency results (1-based)
        lat_old_col = len(key_cols) + 1
        lat_new_col = len(key_cols) + 2
        ratio_col = len(key_cols) + 3
        change_col = len(key_cols) + 4

        # Colour the latency header cells
        ws.cell(row=header_row, column=lat_old_col).fill = _OLD_FILL
        ws.cell(row=header_row, column=lat_old_col).font = Font(bold=True)
        ws.cell(row=header_row, column=lat_new_col).fill = _NEW_FILL
        ws.cell(row=header_row, column=lat_new_col).font = Font(bold=True)

        # --- Data rows (start from row 3) ---
        for row_idx, case_key in enumerate(all_cases, header_row + 1):
            old_row = old_map.get(case_key)
            new_row = new_map.get(case_key)

            # Write config columns
            for col_idx, k in enumerate(key_cols, 1):
                val = case_key[col_idx - 1]
                # Try to convert numeric strings
                try:
                    val = float(val)
                    if val.is_integer():
                        val = int(val)
                except (ValueError, TypeError):
                    pass
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = _THIN_BORDER

            # Old latency
            old_lat = None
            if old_row:
                try:
                    old_lat = float(old_row.get("latency(us)", ""))
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=row_idx, column=lat_old_col, value=old_lat)
            cell.border = _THIN_BORDER
            cell.fill = _OLD_FILL
            if old_lat is not None:
                cell.number_format = "0.000"

            # New latency
            new_lat = None
            if new_row:
                try:
                    new_lat = float(new_row.get("latency(us)", ""))
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=row_idx, column=lat_new_col, value=new_lat)
            cell.border = _THIN_BORDER
            cell.fill = _NEW_FILL
            if new_lat is not None:
                cell.number_format = "0.000"

            # Ratio = new / old  (>1 means new is slower)
            ratio = None
            if old_lat and new_lat and old_lat > 0:
                ratio = new_lat / old_lat
            cell = ws.cell(row=row_idx, column=ratio_col, value=ratio)
            cell.border = _THIN_BORDER
            if ratio is not None:
                cell.number_format = "0.00"

            # Change indicator
            change = ""
            if ratio is not None:
                if ratio < 0.90:
                    change = "↑ faster"
                elif ratio > 1.10:
                    change = "↓ slower"
                else:
                    change = "→ same"
            cell = ws.cell(row=row_idx, column=change_col, value=change)
            cell.border = _THIN_BORDER
            if ratio is not None:
                if ratio < 0.90:
                    cell.fill = _GREEN_FILL
                elif ratio > 1.10:
                    cell.fill = _RED_FILL

        # Auto-fit column widths (approximate)
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for row_idx in range(2, min(ws.max_row + 1, 50)):  # sample first 50 rows
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        # Freeze header row and config columns
        ws.freeze_panes = ws.cell(row=header_row + 1, column=lat_old_col)

    # --- Summary sheet ---
    ws_summary = wb.create_sheet(title="Summary", index=0)
    summary_headers = ["op", "provider", "total_cases", "faster(>10%)", "slower(>10%)", "same",
                       "old_only", "new_only", "avg_ratio", "sheet_link"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    summary_row = 2
    for op, provider in all_keys:
        old_rows = old_data.get((op, provider), [])
        new_rows = new_data.get((op, provider), [])
        if not old_rows and not new_rows:
            continue

        sample = old_rows[0] if old_rows else new_rows[0]
        key_cols = config_keys(sample)
        if new_rows:
            for k in config_keys(new_rows[0]):
                if k not in key_cols:
                    key_cols.append(k)

        old_map = {make_case_key(r, key_cols): r for r in old_rows}
        new_map = {make_case_key(r, key_cols): r for r in new_rows}

        common = set(old_map.keys()) & set(new_map.keys())
        old_only = len(set(old_map.keys()) - set(new_map.keys()))
        new_only = len(set(new_map.keys()) - set(old_map.keys()))

        faster = slower = same = 0
        ratios = []
        for k in common:
            try:
                ol = float(old_map[k].get("latency(us)", ""))
                nl = float(new_map[k].get("latency(us)", ""))
            except (ValueError, TypeError):
                continue
            if ol > 0:
                r = nl / ol
                ratios.append(r)
                if r < 0.90:
                    faster += 1
                elif r > 1.10:
                    slower += 1
                else:
                    same += 1

        total = faster + slower + same
        avg_r = sum(ratios) / len(ratios) if ratios else None

        vals = [op, provider, total, faster, slower, same, old_only, new_only, avg_r]
        for col_idx, v in enumerate(vals, 1):
            cell = ws_summary.cell(row=summary_row, column=col_idx, value=v)
            cell.border = _THIN_BORDER
            if col_idx == 9 and avg_r is not None:
                cell.number_format = "0.00"
        # Highlight summary row
        if avg_r is not None:
            if avg_r < 0.90:
                ws_summary.cell(row=summary_row, column=9).fill = _GREEN_FILL
            elif avg_r > 1.10:
                ws_summary.cell(row=summary_row, column=9).fill = _RED_FILL

        # Sheet link column
        link_col = len(summary_headers)  # last column
        target_sheet = op_sheet_map.get((op, provider))
        if target_sheet:
            link_cell = ws_summary.cell(row=summary_row, column=link_col, value=target_sheet)
            link_cell.hyperlink = f"#'{target_sheet}'!A1"
            link_cell.font = Font(color="0563C1", underline="single")
            link_cell.border = _THIN_BORDER

        summary_row += 1

    # Auto-fit summary columns
    for col_idx in range(1, len(summary_headers) + 1):
        max_len = len(str(ws_summary.cell(row=1, column=col_idx).value or ""))
        for r in range(2, summary_row):
            val = ws_summary.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    wb.save(output_path)
    return len(all_keys)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Compare per-case latency between two reports")
    parser.add_argument("dir1", help="First report directory (auto-detected as old/new by timestamp)")
    parser.add_argument("dir2", help="Second report directory")
    parser.add_argument("-o", "--output", help="Output XLSX path (default: auto-generated)")
    args = parser.parse_args()

    dir1 = os.path.realpath(args.dir1)
    dir2 = os.path.realpath(args.dir2)

    for d in [dir1, dir2]:
        if not os.path.isdir(d):
            print(f"[ERROR] Not a directory: {d}")
            sys.exit(1)

    ts1 = extract_timestamp(dir1)
    ts2 = extract_timestamp(dir2)

    if ts1 and ts2 and ts1 <= ts2:
        old_dir, new_dir = dir1, dir2
    else:
        old_dir, new_dir = dir2, dir1

    old_base = find_intel_base(old_dir)
    new_base = find_intel_base(new_dir)
    if not old_base:
        print(f"[ERROR] No INTEL/<gpu>/ or GPU/<gpu>/ found in {old_dir}")
        sys.exit(1)
    if not new_base:
        print(f"[ERROR] No INTEL/<gpu>/ or GPU/<gpu>/ found in {new_dir}")
        sys.exit(1)

    old_label = make_label(old_dir)
    new_label = make_label(new_dir)

    print(f"Old: {old_dir}  (label: {old_label})")
    print(f"New: {new_dir}  (label: {new_label})")
    print("Loading CSVs...")

    old_data = load_csvs(old_base)
    new_data = load_csvs(new_base)

    print(f"  Old: {sum(len(v) for v in old_data.values())} rows across {len(old_data)} (op, provider) groups")
    print(f"  New: {sum(len(v) for v in new_data.values())} rows across {len(new_data)} (op, provider) groups")

    if args.output:
        output_path = args.output
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        def fmt_ts(ts):
            if ts and len(ts) >= 8:
                return f"{ts[:4]}-{ts[4:6]}-{ts[6:8]}"
            return "unknown"
        output_path = os.path.join(
            script_dir,
            f"latency_compare_{fmt_ts(extract_timestamp(old_dir))}_vs_{fmt_ts(extract_timestamp(new_dir))}.xlsx",
        )

    print(f"Writing {output_path} ...")
    n_sheets = write_xlsx(old_data, new_data, old_label, new_label, output_path)
    print(f"Done. {n_sheets} op sheets + 1 summary sheet.")


if __name__ == "__main__":
    main()
