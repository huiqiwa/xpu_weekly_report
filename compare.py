#!/usr/bin/env python3
"""Compare benchmark results from two architectures into a single XLSX workbook.

Usage:
    python compare.py <report_dir_1> <report_dir_2> [-o output.xlsx]

Each report directory should contain an INTEL/<sku>/ or GPU/<sku>/ subdirectory
with the standard op/provider/csv structure.

Output: one XLSX file with one sheet per op.  Each sheet has:
  - Common test-case configuration columns on the left
  - Two colour-coded compound column groups (one per architecture) on the right,
    each containing: provider | status | latency(us) | mem_bw(GB/s) |
    calc_flops_power(tflops) | MFU(%) | MBU(%)

When an op has multiple providers within one architecture, the successful
provider with the lowest latency is selected for each test case.
"""

import argparse
import copy
import glob
import itertools
import json
import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Peak specs per SKU  (bandwidth in GB/s, compute in TFLOPS)
# ---------------------------------------------------------------------------
_SKU_PEAK_SPECS = {
    "Intel(R) Graphics [0xe211]": {
        "bw_gbs": 456.0,
        "fp32": 12.28,
        "low": 98.5,       # bf16 / fp16
        "int8": 197.0,
    },
    "RTX 5090 D": {
        "bw_gbs": 1792.0,
        "fp32": 96.3,
        "tfloat32": 123.7,
        "low": 246.06208,
        "int8": 480.46,
    },
    "RTX 5090": {
        "bw_gbs": 1792.0,
        "fp32": 96.3,
        "tfloat32": 124.7,
        "low": 237.6,
        "int8": 450.3,
    },
    "RTX PRO 5000": {
        "bw_gbs": 1344.0,
        "fp32": 55.1,
        "tfloat32": 113.6,
        "low": 214.5792,
        "int8": 457.76896,
    },
}

SAGE_ATTENTION_OPS = set()

# CCL (collective communication) ops — use different result columns.
CCL_OPS = {
    "all_reduce", "all_gather", "reduce_scatter", "all_to_all",
    "device2host", "host2device",
}

# Columns that are test *results* rather than case configuration.
# Non-CCL ops: only latency / bandwidth / flops metrics are results.
_METRIC_COLUMNS_DEFAULT = {
    "sku_name", "op_name", "provider",
    "latency(us)", "mem_bw(GB/s)", "calc_flops_power(tflops)",
    "read_bytes(B)", "write_bytes(B)", "io_bytes(B)",
    "calc_flops", "calc_mem_ratio", "kernels",
    "MFU(%)", "MBU(%)", "case_status",
}

# CCL ops: latency + algo/bus bandwidth columns are results.
_METRIC_COLUMNS_CCL = {
    "sku_name", "op_name", "provider",
    "latency(us)",
    "read_bytes(B)", "write_bytes(B)", "io_bytes(B)",
    "mem_bw(GB/s)", "calc_flops", "calc_flops_power(tflops)",
    "calc_mem_ratio", "kernels",
    "algo_bw(GB/s)", "bus_bw(GB/s)",
    "algo_bw_sum(GB/s)", "bus_bw_sum(GB/s)",
    "latency_list(us)", "algo_bw_list(GB/s)", "bus_bw_list(GB/s)",
    "case_status",
}

# Sub-columns shown under each architecture's compound header.
RESULT_SUB_COLUMNS_DEFAULT = [
    "provider", "status", "latency(us)", "mem_bw(GB/s)",
    "calc_flops_power(tflops)", "MFU(%)", "MBU(%)",
]

RESULT_SUB_COLUMNS_CCL = [
    "provider", "status", "latency(us)", "mem_bw(GB/s)",
    "algo_bw(GB/s)", "bus_bw(GB/s)",
    "algo_bw_sum(GB/s)", "bus_bw_sum(GB/s)",
    "latency_list(us)", "algo_bw_list(GB/s)", "bus_bw_list(GB/s)",
]


def _is_ccl_op(op_name):
    return op_name in CCL_OPS


def _metric_columns(op_name):
    return _METRIC_COLUMNS_CCL if _is_ccl_op(op_name) else _METRIC_COLUMNS_DEFAULT


def _result_sub_columns(op_name):
    return RESULT_SUB_COLUMNS_CCL if _is_ccl_op(op_name) else RESULT_SUB_COLUMNS_DEFAULT

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_peak_specs(sku_name):
    for substring, specs in _SKU_PEAK_SPECS.items():
        if substring in sku_name:
            return specs
    print(f"  WARNING: No peak specs for SKU '{sku_name}', using fallback")
    return {"bw_gbs": 1.0, "fp32": 1.0, "low": 1.0, "int8": 1.0}


def _compute_mfu(row, specs, op_name):
    dtype = str(row.get("dtype", "")).lower()
    if op_name in SAGE_ATTENTION_OPS:
        peak = specs["int8"]
    elif dtype in ("tf32", "tfloat32"):
        peak = specs.get("tfloat32", specs["fp32"])
    elif "float32" in dtype or dtype == "fp32":
        peak = specs["fp32"]
    elif dtype == "int8":
        peak = specs["int8"]
    else:
        peak = specs["low"]
    val = row.get("calc_flops_power(tflops)")
    if pd.isna(val) or peak == 0:
        return None
    return round(float(val) / peak * 100, 4)


def _compute_mbu(row, specs):
    val = row.get("mem_bw(GB/s)")
    bw = specs["bw_gbs"]
    if pd.isna(val) or bw == 0:
        return None
    return round(float(val) / bw * 100, 4)


def _normalize_scalar(value):
    """Normalise a scalar value to a canonical string for comparison."""
    if pd.isna(value):
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else f"{value:.12g}".lower()
    text = str(value).strip().lower()
    if text in ("true", "false"):
        return text
    try:
        n = float(text)
    except ValueError:
        return text
    return str(int(n)) if n.is_integer() else f"{n:.12g}"


def _config_columns(df, op_name):
    """Return the ordered list of configuration (non-metric) columns."""
    metrics = _metric_columns(op_name)
    return [c for c in df.columns if c not in metrics and not c.startswith("sycl_ext_")]


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def resolve_report_base(report_dir):
    """Locate the SKU-level directory under a report directory.

    Handles both ``INTEL/<sku>`` and ``GPU/<sku>`` layouts.
    Returns ``(base_path, sku_name)``.
    """
    for vendor in ("INTEL", "GPU"):
        vendor_path = os.path.join(report_dir, vendor)
        if not os.path.isdir(vendor_path):
            continue
        candidates = [
            p for p in sorted(glob.glob(os.path.join(glob.escape(vendor_path), "*")))
            if os.path.isdir(p)
        ]
        if candidates:
            return candidates[0], os.path.basename(candidates[0])
    raise FileNotFoundError(f"No INTEL/* or GPU/* subdirectory in: {report_dir}")


def load_report_frames(base_dir):
    """Load all CSVs grouped as ``{op_name: {provider: DataFrame}}``."""
    pattern = os.path.join(glob.escape(base_dir), "*", "*", "*.csv")
    grouped = {}
    for csv_path in sorted(glob.glob(pattern)):
        if os.path.getsize(csv_path) == 0:
            continue
        try:
            df = pd.read_csv(csv_path)
        except pd.errors.EmptyDataError:
            continue
        if df.empty:
            continue
        rel = os.path.relpath(csv_path, base_dir)
        parts = rel.split(os.sep)
        op_name = parts[0]
        provider = parts[1] if len(parts) > 1 else "unknown"
        grouped.setdefault(op_name, {}).setdefault(provider, []).append(df)
    merged = {}
    for op_name, prov_dict in grouped.items():
        merged[op_name] = {}
        for provider, frames in prov_dict.items():
            df = pd.concat(frames, ignore_index=True, sort=False)
            # Normalise mode / attn_mode for flash_attention:
            # Some providers use "mode" while others use "attn_mode".
            # Unify into "attn_mode" so cross-provider keys match.
            if op_name == "flash_attention":
                if "attn_mode" not in df.columns and "mode" in df.columns:
                    df["attn_mode"] = df["mode"]
                elif "attn_mode" in df.columns and "mode" in df.columns:
                    df["attn_mode"] = df["attn_mode"].fillna(df["mode"])
                if "mode" in df.columns:
                    df.drop(columns=["mode"], inplace=True)
                # Drop config columns that are entirely empty for this provider
                # (they originate from other providers' column superset).
                optional_cols = ["cache_type", "dst_dtype", "qk_compute_dtype", "pv_compute_dtype"]
                for col in optional_cols:
                    if col in df.columns and df[col].isna().all():
                        df.drop(columns=[col], inplace=True)
            merged[op_name][provider] = df
    return merged


# ---------------------------------------------------------------------------
# Per-op comparison logic
# ---------------------------------------------------------------------------

def _extract_result(info, sub_columns):
    """Turn an internal result dict (or None) into a display dict."""
    if info is None:
        d = {c: None for c in sub_columns}
        d["status"] = "FAILED"
        return d
    return {c: info.get(c) for c in sub_columns}


def _safe_round(val, digits):
    if pd.isna(val):
        return None
    return round(float(val), digits)


def process_op(op_name, prov1, prov2, specs1, specs2):
    """Build comparison rows for *op_name* across two architectures.

    Config columns are determined from arch2 (the reference architecture).
    Arch1 providers that do not contain all arch2 config columns are excluded.

    Returns ``{"config_cols": [...], "sub_columns": [...], "rows": [...]}``
    or *None*.
    """
    all_dfs = list(prov1.values()) + list(prov2.values())
    if not all_dfs:
        return None

    sub_columns = _result_sub_columns(op_name)
    is_ccl = _is_ccl_op(op_name)

    # Determine config columns from arch2 (reference).  If arch2 has no data
    # for this op, fall back to arch1's intersection.
    if prov2:
        arch2_cfg_sets = [frozenset(_config_columns(df, op_name)) for df in prov2.values()]
        ref_cfg = arch2_cfg_sets[0]
        for s in arch2_cfg_sets[1:]:
            ref_cfg &= s
        # Preserve column order from the first arch2 DataFrame.
        ref_df = next(iter(prov2.values()))
        config_cols = [c for c in ref_df.columns if c in ref_cfg]
    elif prov1:
        arch1_cfg_sets = [frozenset(_config_columns(df, op_name)) for df in prov1.values()]
        ref_cfg = arch1_cfg_sets[0]
        for s in arch1_cfg_sets[1:]:
            ref_cfg &= s
        ref_df = next(iter(prov1.values()))
        config_cols = [c for c in ref_df.columns if c in ref_cfg]
    else:
        return None

    # Also include config columns that exist in arch1 but not arch2,
    # so that e.g. block_size from Intel data is not lost.
    # Only include columns present in ALL arch1 providers (intersection).
    if prov1 and prov2:
        arch1_cfg_sets = [frozenset(_config_columns(df, op_name)) for df in prov1.values()]
        arch1_cfg_intersection = arch1_cfg_sets[0]
        for s in arch1_cfg_sets[1:]:
            arch1_cfg_intersection &= s
        existing = frozenset(config_cols)
        arch1_ref_df = next(iter(prov1.values()))
        extra = [c for c in arch1_ref_df.columns if c in arch1_cfg_intersection and c not in existing]
        if extra:
            config_cols = config_cols + extra
            ref_cfg = ref_cfg | frozenset(extra)

    if not config_cols:
        return None

    # Filter arch1 providers: keep only those whose columns contain all
    # reference config columns.
    filtered_prov1 = {}
    for pname, df in prov1.items():
        prov_cfg = frozenset(_config_columns(df, op_name))
        if ref_cfg <= prov_cfg:
            filtered_prov1[pname] = df
    if prov1 and not filtered_prov1:
        # No qualifying arch1 providers — keep the most complete one as
        # fallback so that the op still appears (rows will show FAILED).
        best_prov = max(prov1, key=lambda p: len(set(prov1[p].columns) & ref_cfg))
        filtered_prov1[best_prov] = prov1[best_prov]

    def _case_key(row):
        return tuple(_normalize_scalar(row.get(c)) for c in config_cols)

    def _pick_best(prov_dict, specs):
        """For each test case key, pick the best provider row."""
        cases = {}  # key -> [(provider, row), ...]
        for provider, df in prov_dict.items():
            for _, row in df.iterrows():
                k = _case_key(row)
                cases.setdefault(k, []).append((provider, row))

        best = {}
        for k, entries in cases.items():
            ok = [
                (p, r) for p, r in entries
                if not pd.isna(r.get("latency(us)"))
            ]
            if ok:
                provider, row = min(ok, key=lambda x: float(x[1]["latency(us)"]))
                status = "SUCCESS"
            else:
                provider, row = entries[0]
                status = "FAILED"

            result = {
                "config": {c: row.get(c) for c in config_cols},
                "provider": provider,
                "status": status,
                "latency(us)": _safe_round(row.get("latency(us)"), 3),
            }

            if is_ccl:
                for col in [
                    "algo_bw(GB/s)", "bus_bw(GB/s)",
                    "algo_bw_sum(GB/s)", "bus_bw_sum(GB/s)",
                ]:
                    result[col] = _safe_round(row.get(col), 3)
                for col in [
                    "latency_list(us)", "algo_bw_list(GB/s)", "bus_bw_list(GB/s)",
                ]:
                    val = row.get(col)
                    result[col] = val if not pd.isna(val) else None
            else:
                result["mem_bw(GB/s)"] = _safe_round(row.get("mem_bw(GB/s)"), 3)
                result["calc_flops_power(tflops)"] = _safe_round(
                    row.get("calc_flops_power(tflops)"), 4
                )
                result["MFU(%)"] = _compute_mfu(row, specs, op_name)
                result["MBU(%)"] = _compute_mbu(row, specs)

            best[k] = result
        return best

    best1 = _pick_best(filtered_prov1, specs1) if filtered_prov1 else {}
    best2 = _pick_best(prov2, specs2) if prov2 else {}

    all_keys = sorted(set(best1) | set(best2))
    rows = []
    for k in all_keys:
        r1 = best1.get(k)
        r2 = best2.get(k)
        config = (r1 or r2)["config"]
        rows.append({
            "config": config,
            "arch1": _extract_result(r1, sub_columns),
            "arch2": _extract_result(r2, sub_columns),
        })
    return {
        "config_cols": config_cols,
        "sub_columns": sub_columns,
        "rows": rows,
    } if rows else None


# ---------------------------------------------------------------------------
# Workload case loading (from JSON definitions)
# ---------------------------------------------------------------------------

def _value_options(key, value):
    parts = key.split(".")
    if len(parts) > 1:
        if isinstance(value, list):
            if value and isinstance(value[0], list):
                return [dict(zip(parts, item)) for item in value]
            if len(value) == len(parts):
                return [dict(zip(parts, value))]
        return [dict(zip(parts, [value]))]
    if isinstance(value, list):
        return [{key: item} for item in value]
    return [{key: value}]


def _expand_case(case_def):
    option_groups = [_value_options(k, v) for k, v in case_def.items()]
    expanded = []
    for combination in itertools.product(*option_groups):
        merged = {}
        for item in combination:
            merged.update(item)
        expanded.append(merged)
    return expanded


def load_workload_cases(workloads_dir):
    """Load workload case definitions from JSON files.

    Returns ``{op_name: [dict, ...]}``.  Each dict represents one
    fully-expanded test case.
    """
    workload_cases = {}
    json_files = sorted(
        glob.glob(os.path.join(glob.escape(workloads_dir), "**", "*.json"), recursive=True)
    )
    for json_path in json_files:
        with open(json_path, "r", encoding="utf-8") as f:
            payload = json.load(f)

        if isinstance(payload, dict) and "cases" in payload:
            op_name = os.path.splitext(os.path.basename(json_path))[0]
            workload_cases.setdefault(op_name, [])
            for case_def in payload["cases"]:
                workload_cases[op_name].extend(_expand_case(case_def))
            continue

        if not isinstance(payload, dict):
            continue
        for op_name, case_defs in payload.items():
            if not isinstance(case_defs, list):
                continue
            workload_cases.setdefault(op_name, [])
            for case_def in case_defs:
                workload_cases[op_name].extend(_expand_case(case_def))

    return workload_cases


# Ordered list of ops to include (and their sheet order).
ORDERED_OPS = [
    "scale_dynamic_quant",
    "head_rms_norm",
    "head_rms_norm_dynamic_quant",
    "add_rms_norm_dynamic_quant",
    "rotary_embedding",
    "store_kv_cache",
    "dequant_kv_cache",
    "flash_attention",
    "moe_gating_gemm",
    "quant_matmul",
    "moe_quant_group_gemm",
    "moe_softmax_topk",
    "moe_scatter_dynamic_quant",
    "moe_swiglu_dynamic_quant",
    "swiglu_dynamic_quant",
    "moe_gather",
    "moe_quant_group_gemm_combine",
    "quant_group_gemm_reduce_sum",
    "gemm",
    "gelu",
    "silu",
    "embedding",
    "gather",
    "index_add",
    "index_select",
    "scatter",
    "add",
    "cast",
    "mul",
    "sub",
    "layer_norm",
    "rms_norm",
    "softmax",
    "reduce_max",
    "reduce_min",
    "reduce_sum",
    "topk",
    "cos",
    "div",
    "exp",
    "log",
    "sin",
    "sqrt",
    "device2device",
    "device2host",
    "host2device",
    "all_reduce",
    "all_gather",
    "reduce_scatter",
    "all_to_all",
]

# ---------------------------------------------------------------------------
# XLSX output
# ---------------------------------------------------------------------------

def _auto_fit_columns(ws, max_width=80):
    for col_cells in ws.columns:
        max_len = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 4, max_width)


def _safe_ratio(a, b):
    """Return a/b rounded to 4 digits, or None if not computable."""
    if a is None or b is None:
        return None
    try:
        a_f, b_f = float(a), float(b)
    except (TypeError, ValueError):
        return None
    if b_f == 0:
        return None
    return round(a_f / b_f, 4)


def write_comparison_xlsx(path, all_ops_data, sku1, sku2, workload_cases=None):
    wb = Workbook()
    wb.remove(wb.active)

    wl_cases = workload_cases or {}

    # ---- Styles ----
    # Config header
    cfg_hdr_fill = PatternFill("solid", fgColor="D9C6A5")  # warm tan
    cfg_hdr_font = Font(bold=True, size=11, color="3B2F1E")
    # Config data
    cfg_data_fill = PatternFill("solid", fgColor="FFF8EE")  # light cream
    # Arch compound headers
    arch1_hdr_fill = PatternFill("solid", fgColor="4472C4")    # blue
    arch2_hdr_fill = PatternFill("solid", fgColor="70AD47")    # green
    arch_hdr_font = Font(bold=True, size=12, color="FFFFFF")
    sub_hdr_font = Font(bold=True, size=10)
    arch1_sub_fill = PatternFill("solid", fgColor="D6E4F0")
    arch2_sub_fill = PatternFill("solid", fgColor="E2EFDA")
    arch1_data_fill = PatternFill("solid", fgColor="DCE6F1")
    arch2_data_fill = PatternFill("solid", fgColor="E8F0DE")
    # Ratio columns
    ratio_hdr_fill = PatternFill("solid", fgColor="FFF2CC")  # warm yellow
    ratio_hdr_font = Font(bold=True, size=10, color="7F6000")
    ratio_data_fill = PatternFill("solid", fgColor="FFFBE6")
    # Summary
    sum_hdr_fill = PatternFill("solid", fgColor="4472C4")
    sum_hdr_font = Font(bold=True, size=11, color="FFFFFF")
    sum_data_fill_even = PatternFill("solid", fgColor="F2F7FB")
    sum_data_fill_odd = PatternFill("solid", fgColor="FFFFFF")
    thin = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    # ---- Summary sheet ----
    sum_ws = wb.create_sheet("Summary")
    sum_headers = ["op_name", "total_cases", f"failed ({sku1})", f"failed ({sku2})", "sheet_link"]
    for ci, h in enumerate(sum_headers, 1):
        c = sum_ws.cell(1, ci, h)
        c.font = sum_hdr_font
        c.fill = sum_hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin
    sum_row = 2

    # ---- Per-op sheets ----
    for op_name in ORDERED_OPS:
        if op_name not in all_ops_data:
            continue
        data = all_ops_data[op_name]
        cc = data["config_cols"]
        sub_cols = data["sub_columns"]
        rows = data["rows"]
        n_cfg = len(cc)
        n_sub = len(sub_cols)
        is_ccl = _is_ccl_op(op_name)

        sheet = op_name[:31]
        ws = wb.create_sheet(sheet)

        # Column ranges (1-based)
        a1_start = n_cfg + 1
        a1_end = a1_start + n_sub - 1
        a2_start = a1_end + 1
        a2_end = a2_start + n_sub - 1

        # Ratio columns for non-CCL ops; bus_bw ratio for CCL ops
        ratio_cols = []
        if is_ccl:
            ratio_cols = ["bus_bw ratio"]
        else:
            ratio_cols = ["MFU ratio", "MBU ratio"]
        n_ratio = len(ratio_cols)
        ratio_start = a2_end + 1
        total_cols = ratio_start + n_ratio - 1

        # -- Row 1: Back to Summary link --
        ws.cell(1, 1, "← Back to Summary").hyperlink = "#Summary!A1"
        ws.cell(1, 1).style = "Hyperlink"

        # -- Row 2: Arch group name labels (decorative merged row) --
        # Leave config area blank on row 2
        for i in range(n_cfg):
            c = ws.cell(2, i + 1)
            c.fill = cfg_hdr_fill
            c.border = thin

        # Arch-1 merged header (row 2)
        if n_sub > 1:
            ws.merge_cells(
                start_row=2, start_column=a1_start,
                end_row=2, end_column=a1_end,
            )
        c = ws.cell(2, a1_start, sku1)
        c.font = arch_hdr_font
        c.fill = arch1_hdr_fill
        c.alignment = Alignment(horizontal="center")
        for col in range(a1_start, a1_end + 1):
            ws.cell(2, col).fill = arch1_hdr_fill
            ws.cell(2, col).border = thin

        # Arch-2 merged header (row 2)
        if n_sub > 1:
            ws.merge_cells(
                start_row=2, start_column=a2_start,
                end_row=2, end_column=a2_end,
            )
        c = ws.cell(2, a2_start, sku2)
        c.font = arch_hdr_font
        c.fill = arch2_hdr_fill
        c.alignment = Alignment(horizontal="center")
        for col in range(a2_start, a2_end + 1):
            ws.cell(2, col).fill = arch2_hdr_fill
            ws.cell(2, col).border = thin

        # Comparison merged header (row 2)
        if n_ratio > 1:
            ws.merge_cells(
                start_row=2, start_column=ratio_start,
                end_row=2, end_column=ratio_start + n_ratio - 1,
            )
        rc = ws.cell(2, ratio_start, "Comparison")
        rc.font = Font(bold=True, size=12, color="7F6000")
        rc.fill = ratio_hdr_fill
        rc.alignment = Alignment(horizontal="center")
        for col in range(ratio_start, ratio_start + n_ratio):
            ws.cell(2, col).fill = ratio_hdr_fill
            ws.cell(2, col).border = thin

        # -- Row 3: Column headers (table header row) --
        # Build unique header names for the table
        all_headers = []
        for col_name in cc:
            all_headers.append(col_name)
        for sc in sub_cols:
            all_headers.append(sc)
        for sc in sub_cols:
            all_headers.append(sc + "_2")
        for rc_name in ratio_cols:
            all_headers.append(rc_name)

        for ci, h in enumerate(all_headers):
            col_idx = ci + 1
            c = ws.cell(3, col_idx, h)
            c.font = sub_hdr_font
            c.border = thin
            if col_idx <= n_cfg:
                c.font = cfg_hdr_font
                c.fill = cfg_hdr_fill
                c.alignment = Alignment(vertical="center", horizontal="center")
            elif col_idx <= a1_end:
                c.fill = arch1_sub_fill
            elif col_idx <= a2_end:
                c.fill = arch2_sub_fill
            else:
                c.font = ratio_hdr_font
                c.fill = ratio_hdr_fill

        # -- Data rows (starting at row 4) --
        for ri, rd in enumerate(rows):
            er = ri + 4
            cfg = rd["config"]
            for ci, col in enumerate(cc):
                cell = ws.cell(er, ci + 1, cfg.get(col))
                cell.fill = cfg_data_fill
                cell.border = thin
            for si, sc in enumerate(sub_cols):
                v1 = rd["arch1"].get(sc)
                c1 = ws.cell(er, a1_start + si, v1)
                c1.fill = arch1_data_fill
                c1.border = thin
                v2 = rd["arch2"].get(sc)
                c2 = ws.cell(er, a2_start + si, v2)
                c2.fill = arch2_data_fill
                c2.border = thin

            # Ratio data columns
            if is_ccl:
                bw1 = rd["arch1"].get("bus_bw(GB/s)")
                bw2 = rd["arch2"].get("bus_bw(GB/s)")
                c = ws.cell(er, ratio_start, _safe_ratio(bw1, bw2))
                c.fill = ratio_data_fill
                c.border = thin
            else:
                mfu1 = rd["arch1"].get("MFU(%)")
                mfu2 = rd["arch2"].get("MFU(%)")
                mbu1 = rd["arch1"].get("MBU(%)")
                mbu2 = rd["arch2"].get("MBU(%)")
                c = ws.cell(er, ratio_start, _safe_ratio(mfu1, mfu2))
                c.fill = ratio_data_fill
                c.border = thin
                c = ws.cell(er, ratio_start + 1, _safe_ratio(mbu1, mbu2))
                c.fill = ratio_data_fill
                c.border = thin

        # Add Excel Table
        last_data_row = 3 + len(rows)
        table_ref = f"A3:{get_column_letter(total_cols)}{last_data_row}"
        tbl = Table(displayName=re.sub(r'[^A-Za-z0-9_]', '_', op_name),
                    ref=table_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tbl)

        # Freeze panes so headers stay visible
        ws.freeze_panes = "A4"
        _auto_fit_columns(ws)

        # Summary entry
        expected = len(wl_cases.get(op_name, []))
        total_cases = expected if expected else len(rows)
        failed1 = sum(1 for r in rows if r["arch1"].get("status") == "FAILED")
        failed2 = sum(1 for r in rows if r["arch2"].get("status") == "FAILED")
        row_fill = sum_data_fill_even if (sum_row % 2 == 0) else sum_data_fill_odd
        c = sum_ws.cell(sum_row, 1, op_name)
        c.border = thin
        c.fill = row_fill
        c = sum_ws.cell(sum_row, 2, total_cases)
        c.border = thin
        c.fill = row_fill
        c.alignment = Alignment(horizontal="center")
        c = sum_ws.cell(sum_row, 3, failed1)
        c.border = thin
        c.fill = row_fill
        c.alignment = Alignment(horizontal="center")
        c = sum_ws.cell(sum_row, 4, failed2)
        c.border = thin
        c.fill = row_fill
        c.alignment = Alignment(horizontal="center")
        link_cell = sum_ws.cell(sum_row, 5, sheet)
        link_cell.hyperlink = f"#{sheet}!A1"
        link_cell.style = "Hyperlink"
        link_cell.border = thin
        link_cell.fill = row_fill
        sum_row += 1

        print(f"  Sheet: {sheet} ({len(rows)} cases)")

    # Add Excel Table to Summary sheet
    if sum_row > 2:
        sum_tbl_ref = f"A1:E{sum_row - 1}"
        sum_tbl = Table(displayName="Summary", ref=sum_tbl_ref)
        sum_tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        sum_ws.add_table(sum_tbl)

    _auto_fit_columns(sum_ws)
    wb.save(path)
    print(f"\nDone! Output: {path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    _parent_dir = os.path.dirname(_script_dir)
    default_workloads = os.path.join(_parent_dir, "xpu-perf", "micro_perf", "workloads")
    p = argparse.ArgumentParser(
        description="Compare benchmark results from two architectures."
    )
    p.add_argument("dir1", help="First report directory (e.g. Intel reports)")
    p.add_argument("dir2", help="Second report directory (e.g. NV reports)")
    p.add_argument(
        "-o", "--output", default=None,
        help="Output XLSX file path (default: comparison_{arch1}_vs_{arch2}_{timestamp}.xlsx)",
    )
    p.add_argument(
        "--workloads-dir", default=default_workloads,
        help="Workloads JSON directory (default: xpu-perf/micro_perf/workloads)",
    )
    return p.parse_args()


def main():
    args = parse_args()

    base1, sku1 = resolve_report_base(args.dir1)
    base2, sku2 = resolve_report_base(args.dir2)
    specs1 = _get_peak_specs(sku1)
    specs2 = _get_peak_specs(sku2)

    print(f"Arch 1: {sku1}  (BW={specs1['bw_gbs']} GB/s)")
    print(f"Arch 2: {sku2}  (BW={specs2['bw_gbs']} GB/s)")

    if args.output is None:
        def _sanitize(name):
            return re.sub(r'[^\w\-]+', '_', name).strip('_')
        ts = datetime.now().strftime('%Y-%m-%d')
        args.output = f"comparison_{_sanitize(sku1)}_vs_{_sanitize(sku2)}_{ts}.xlsx"

    frames1 = load_report_frames(base1)
    frames2 = load_report_frames(base2)

    wl_cases = {}
    if os.path.isdir(args.workloads_dir):
        wl_cases = load_workload_cases(args.workloads_dir)
        wl_cases = {op: cases for op, cases in wl_cases.items() if op in ORDERED_OPS}
        print(f"Workload definitions: {sum(len(v) for v in wl_cases.values())} cases across {len(wl_cases)} ops")
    else:
        print(f"WARNING: workloads dir not found: {args.workloads_dir}")

    all_ops = [op for op in ORDERED_OPS if op in frames1 or op in frames2]
    results = {}
    for op in all_ops:
        r = process_op(
            op,
            frames1.get(op, {}),
            frames2.get(op, {}),
            specs1, specs2,
        )
        if r:
            results[op] = r

    # Append workload-defined cases missing from both architectures.
    _append_missing_workload_cases(results, wl_cases)

    # Warn about cases that have data but are NOT defined in workload JSON.
    _warn_extra_cases(results, wl_cases)

    write_comparison_xlsx(args.output, results, sku1, sku2, wl_cases)


def _append_missing_workload_cases(results, wl_cases):
    """Add rows for workload-defined cases missing from both architectures,
    and duplicate existing rows to match JSON-defined repetitions."""
    if not wl_cases:
        return
    for op_name, op_data in results.items():
        expected = wl_cases.get(op_name)
        if not expected:
            continue
        config_cols = op_data["config_cols"]
        sub_columns = op_data["sub_columns"]
        rows = op_data["rows"]

        # Determine which JSON columns overlap with config columns.
        json_keys_cols = set()
        for case_def in expected:
            json_keys_cols.update(case_def.keys())
        match_cols = [c for c in config_cols if c in json_keys_cols]
        if not match_cols:
            continue

        # Build lookup from key -> existing row.
        existing_by_key = {}
        for row in rows:
            cfg = row["config"]
            key = tuple(_normalize_scalar(cfg.get(c)) for c in match_cols)
            existing_by_key.setdefault(key, row)

        # Rebuild rows from workload definitions, preserving JSON order and
        # duplicates.  Each JSON case maps to either the existing data row or
        # a FAILED/FAILED placeholder.
        failed_result = {c: None for c in sub_columns}
        failed_result["status"] = "FAILED"

        new_rows = []
        seen_keys = set()  # track keys already emitted from JSON
        added_missing = 0
        added_dup = 0
        for case_def in expected:
            key = tuple(_normalize_scalar(case_def.get(c)) for c in match_cols)
            existing = existing_by_key.get(key)
            if existing is not None:
                if key not in seen_keys:
                    # First appearance — use original row
                    new_rows.append(existing)
                else:
                    # Duplicate in JSON — copy existing row
                    new_rows.append(copy.deepcopy(existing))
                    added_dup += 1
            else:
                # Case defined in JSON but not in any data
                config = {}
                for c in config_cols:
                    config[c] = case_def.get(c)
                new_rows.append({
                    "config": config,
                    "arch1": dict(failed_result),
                    "arch2": dict(failed_result),
                })
                added_missing += 1
            seen_keys.add(key)

        # Append data-only rows (exist in data but not in JSON) at the end.
        json_keys_all = set()
        for case_def in expected:
            key = tuple(_normalize_scalar(case_def.get(c)) for c in match_cols)
            json_keys_all.add(key)
        for row in rows:
            cfg = row["config"]
            key = tuple(_normalize_scalar(cfg.get(c)) for c in match_cols)
            if key not in json_keys_all:
                new_rows.append(row)

        op_data["rows"] = new_rows
        msgs = []
        if added_missing:
            msgs.append(f"{added_missing} missing (both FAILED)")
        if added_dup:
            msgs.append(f"{added_dup} duplicated")
        if msgs:
            print(f"  + {op_name}: {', '.join(msgs)}")


def _warn_extra_cases(results, wl_cases):
    """Print warnings for ops whose CSV data contains cases not in JSON."""
    if not wl_cases:
        return
    warned = False
    for op_name, op_data in results.items():
        expected = wl_cases.get(op_name)
        if not expected:
            continue
        config_cols = op_data["config_cols"]
        rows = op_data["rows"]

        # Build set of normalised JSON case keys (using only config columns
        # that exist in both JSON definition and data).
        json_keys_cols = set()
        for case_def in expected:
            json_keys_cols.update(case_def.keys())
        match_cols = [c for c in config_cols if c in json_keys_cols]
        if not match_cols:
            continue

        json_key_set = set()
        for case_def in expected:
            key = tuple(_normalize_scalar(case_def.get(c)) for c in match_cols)
            json_key_set.add(key)

        extra_rows = []
        for row in rows:
            cfg = row["config"]
            key = tuple(_normalize_scalar(cfg.get(c)) for c in match_cols)
            if key not in json_key_set:
                extra_rows.append(cfg)

        if extra_rows:
            if not warned:
                print("\n--- WARNING: cases in data but NOT in workload JSON ---")
                warned = True
            print(f"\n  [{op_name}] {len(extra_rows)} extra case(s)  (match_cols={match_cols}):")
            for cfg in extra_rows:
                vals = ", ".join(f"{c}={cfg.get(c)}" for c in config_cols)
                print(f"    - {vals}")

    if warned:
        print()


if __name__ == "__main__":
    main()
